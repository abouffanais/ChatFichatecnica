"""
actualizar_stock.py
===================
Hace TODO automatico:
  1. Abre el Excel y refresca la conexion ODBC con SAP
  2. Convierte los datos a stock.json
  3. Guarda el archivo listo para subir a GitHub

REQUISITOS:
  pip install openpyxl pywin32
  (pywin32 solo funciona en Windows)

USO:
  - Doble clic en AUTO_STOCK_GITHUB.bat  (o python actualizar_stock.py)
  - Debes estar en WiFi de KC o conectado a VPN con FortiToken
"""

import os
import sys
import json
import time
import openpyxl
from datetime import datetime

# Forzar salida UTF-8 para evitar errores con caracteres especiales
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# --- CONFIGURACION -------------------------------------------------------
EXCEL_PATH  = r"C:\Users\abouffanais\Documents\Claude\Projects\ChatbotCyber\Stock_disponible.xlsx"
HOJA        = "Consulta1"
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "netlify", "functions", "stock.json")
# -------------------------------------------------------------------------


def refrescar_excel(excel_path):
    """
    Refresca el Excel via COM sin cerrar otros libros abiertos.
    - Si Excel ya estaba abierto: usa esa instancia, no la cierra al terminar
    - Si Excel no estaba abierto: abre una instancia nueva y la cierra al terminar
    - Si el archivo ya estaba abierto: no lo cierra al terminar
    """
    try:
        import win32com.client
        abs_path = os.path.abspath(excel_path)

        # Intentar conectarse a Excel ya abierto
        excel_ya_abierto = False
        wb_ya_abierto    = False
        excel = None
        wb    = None

        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
            excel_ya_abierto = True
            print("  Usando instancia de Excel ya abierta...")
        except Exception:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel_ya_abierto = False
            print("  Abriendo nueva instancia de Excel...")

        excel.DisplayAlerts = False

        # Verificar si el archivo ya esta abierto en Excel
        for w in excel.Workbooks:
            if w.FullName.lower() == abs_path.lower():
                wb = w
                wb_ya_abierto = True
                print("  Archivo ya abierto en Excel, refrescando...")
                break

        if wb is None:
            wb = excel.Workbooks.Open(abs_path)
            wb_ya_abierto = False

        # Refrescar conexiones ODBC
        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()
        time.sleep(3)
        wb.Save()

        # Cerrar solo lo que abrimos nosotros
        if not wb_ya_abierto:
            wb.Close(False)
        if not excel_ya_abierto:
            excel.Quit()

        print("  Excel refrescado y guardado [OK]")
        return True

    except ImportError:
        print("  [AVISO] pywin32 no instalado - saltando refresco automatico")
        print("  Instala con: pip install pywin32")
        print("  Usando el Excel tal como esta guardado...\n")
        return False
    except Exception as e:
        print(f"  [ERROR] al refrescar Excel: {e}")
        print("  Usando el Excel tal como esta guardado...\n")
        return False


def convertir_excel_a_json(excel_path, hoja):
    print(f"  Leyendo: {excel_path}  (hoja: {hoja})")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[hoja]

    catalogo    = {}
    total_filas = ws.max_row - 1
    procesadas  = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku, nombre, categoria, marca, obsoleto, \
        cod_almacen, en_stock, comprometido, disponible, nombre_almacen = row

        if sku is None:
            continue

        sku_str      = str(sku).strip()
        disponible   = int(disponible)   if disponible   is not None else 0
        en_stock     = int(en_stock)     if en_stock     is not None else 0
        comprometido = int(comprometido) if comprometido is not None else 0

        if sku_str not in catalogo:
            catalogo[sku_str] = {
                "sku"      : sku_str,
                "nombre"   : str(nombre).strip()    if nombre    else "",
                "categoria": str(categoria).strip() if categoria else "",
                "marca"    : str(marca).strip()     if marca     else "",
                "obsoleto" : True if str(obsoleto).strip().upper() == "S" else False,
                "bodegas"  : [],
                "total_stock_disponible": 0,
                "total_en_stock"        : 0,
                "total_comprometido"    : 0
            }

        if en_stock != 0 or disponible != 0 or comprometido != 0:
            catalogo[sku_str]["bodegas"].append({
                "codigo"      : str(cod_almacen).strip()    if cod_almacen    else "",
                "nombre"      : str(nombre_almacen).strip() if nombre_almacen else "",
                "en_stock"    : en_stock,
                "comprometido": comprometido,
                "disponible"  : disponible
            })

        catalogo[sku_str]["total_stock_disponible"] += disponible
        catalogo[sku_str]["total_en_stock"]          += en_stock
        catalogo[sku_str]["total_comprometido"]      += comprometido

        procesadas += 1
        if procesadas % 5000 == 0:
            print(f"  Procesando... {procesadas}/{total_filas} filas")

    wb.close()
    return catalogo


def main():
    print("=" * 55)
    print("  Kitchen Center - Actualizacion de Stock")
    print("=" * 55)
    print(f"\n  {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    if not os.path.exists(EXCEL_PATH):
        print(f"\n  ERROR: No se encontro el archivo:\n  {EXCEL_PATH}")
        print("\n  Ajusta la variable EXCEL_PATH en este script.")
        return

    # Paso 1: Refrescar Excel con SAP
    refrescar_excel(EXCEL_PATH)

    # Paso 2: Convertir a JSON
    print("\n  Convirtiendo a JSON...")
    catalogo = convertir_excel_a_json(EXCEL_PATH, HOJA)

    skus_total     = len(catalogo)
    skus_con_stock = sum(1 for v in catalogo.values() if v["total_stock_disponible"] > 0)
    print(f"  SKUs procesados : {skus_total:,}")
    print(f"  Con stock > 0   : {skus_con_stock:,}")

    # Paso 3: Guardar JSON
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    output = {
        "metadata": {
            "actualizado_el": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "total_skus"    : skus_total,
            "skus_con_stock": skus_con_stock,
            "fuente"        : "SAP Business One via ODBC"
        },
        "productos": catalogo
    }
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    size_kb = os.path.getsize(OUTPUT_FILE) / 1024
    print(f"\n  [OK] Archivo generado: {OUTPUT_FILE}")
    print(f"  Tamano: {size_kb:.0f} KB")
    print("=" * 55)


if __name__ == "__main__":
    main()
